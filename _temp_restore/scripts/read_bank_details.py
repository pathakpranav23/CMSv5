import os
import sys
import json
from typing import List, Dict, Any

try:
    import openpyxl
except ImportError as e:
    print("ERROR: openpyxl not installed. Install with 'pip install openpyxl'.")
    sys.exit(1)


def normalize_key(s: str) -> str:
    if s is None:
        return ""
    return "".join(ch.lower() for ch in str(s) if ch.isalnum())


def load_rows(path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    # Prefer first sheet explicitly to avoid surprises
    ws = wb[wb.sheetnames[0]]

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers: List[str] = [c.value if c.value is not None else f"col_{i+1}" for i, c in enumerate(header_cells)]
    norm_headers = [normalize_key(h) for h in headers]

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
        rows.append(d)

    return {
        "sheetnames": wb.sheetnames,
        "headers": headers,
        "norm_headers": norm_headers,
        "rows": rows,
    }


def guess_keys(headers: List[str]) -> Dict[str, str]:
    # Map logical fields to actual header names by fuzzy matching
    norm_map = {normalize_key(h): h for h in headers}
    def find(*candidates: str) -> str:
        for cand in candidates:
            # exact normalized match
            if cand in norm_map:
                return norm_map[cand]
        # contains matching
        for norm, orig in norm_map.items():
            if any(cand in norm for cand in candidates):
                return orig
        return ""

    return {
        "program": find("program", "course", "programme"),
        "bank_name": find("bankname", "bank"),
        "account_name": find("accountname", "accountholder", "beneficiaryname", "payee"),
        "account_number": find("accountnumber", "accountno", "accno"),
        "ifsc": find("ifsc", "ifsccode"),
        "branch": find("branch", "bankbranch"),
        "upi_vpa": find("upivpa", "upiid", "vpa"),
        "payee_display": find("payeedisplay", "payeenickname", "displayname"),
        "gstin": find("gstin", "gst"),
        "pan": find("pan"),
    }


def summarize_by_program(rows: List[Dict[str, Any]], keys: Dict[str, str]) -> Dict[str, Dict[str, Any]]:
    prog_key = keys.get("program", "")
    summary: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        prog = str(r.get(prog_key, "")).strip() if prog_key else ""
        if not prog:
            # Skip rows without program/course info
            continue
        entry = summary.setdefault(prog, {})
        for k, hdr in keys.items():
            if not hdr:
                continue
            val = r.get(hdr)
            if val is not None and str(val).strip():
                entry[k] = str(val).strip()
    return summary


def main():
    # Default path relative to project root
    default_rel = os.path.join("DATA FOR IMPORT EXPORT", "SBPET All Bank Detail.xlsx")
    path = sys.argv[1] if len(sys.argv) > 1 else default_rel
    if not os.path.isabs(path):
        path = os.path.abspath(path)
    if not os.path.exists(path):
        print(json.dumps({"error": f"File not found: {path}"}))
        sys.exit(1)

    data = load_rows(path)
    keys = guess_keys(data["headers"])
    summary = summarize_by_program(data["rows"], keys)

    output = {
        "file": path,
        "sheets": data["sheetnames"],
        "headers": data["headers"],
        "guessed_keys": keys,
        "row_count": len(data["rows"]),
        "program_summary": summary,
        "sample_rows": data["rows"][:10],
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()