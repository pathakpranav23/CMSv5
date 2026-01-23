import re
import sys
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd
from sqlalchemy import select

from cms_app import create_app, db
from cms_app.models import Program, SubjectType, Subject, CreditStructure


# Synonyms to normalize headers from university Excel files
HEADER_MAP: Dict[str, List[str]] = {
    "subject_name": [
        "subject",
        "subject name",
        "course",
        "paper",
        "title",
        "name",
        "sub name",
    ],
    "subject_code": [
        "code",
        "subject code",
        "course code",
        "code no",
        "subject id",
        "sub code",
    ],
    "paper_code": [
        "paper code",
        "paper id",
        "paper no",
    ],
    "subject_type": [
        "type",
        "subject type",
        "category",
        "course type",
        "type code",
    ],
    "semester": [
        "semester",
        "sem",
        "semester no",
        "sem no",
    ],
    "theory_credits": [
        "theory credits",
        "theory",
        "th credits",
        "credits theory",
        "th",
        "credit",
    ],
    "practical_credits": [
        "practical credits",
        "practical",
        "pr credits",
        "credits practical",
        "pr",
    ],
    "total_credits": [
        "total credits",
        "credits",
        "credit points",
        "credits total",
        "points",
    ],
}


def normalize_headers(headers: List[Any]) -> Dict[int, str]:
    colmap: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        key = str(h or "").strip().lower()
        if not key:
            continue
        for target, synonyms in HEADER_MAP.items():
            if key == target:
                colmap[idx] = target
                break
            if key in synonyms:
                colmap[idx] = target
                break
    return colmap


def cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # Avoid trailing .0 for integer-like values
        if abs(v - int(v)) < 1e-9:
            return str(int(v))
        return str(v)
    return str(v)


def detect_program_and_semester(path: str) -> Tuple[str, int]:
    # Enhance detection to support BCom (English) and BCom (Gujarati) variants
    s = (path or "").lower()
    # Normalize separators and punctuation for easier matching
    s = s.replace("_", " ").replace("-", " ")

    # Program detection
    if "bca" in s:
        program_name = "BCA"
    elif "bba" in s:
        program_name = "BBA"
    elif ("bcom" in s) or ("b.com" in s) or ("b com" in s):
        if ("eng" in s) or ("english" in s):
            program_name = "BCom (English)"
        elif ("guj" in s) or ("gujarati" in s):
            program_name = "BCom (Gujarati)"
        else:
            program_name = "BCOM"
    else:
        # Fallback to BCA
        program_name = "BCA"

    # Semester detection: look for 'sem' or 'semester' followed by a number
    m_sem = re.search(r"\b(?:sem(?:ester)?)\b[\s_-]*([0-9]{1,2})", s, flags=re.IGNORECASE)
    if not m_sem:
        # Also support patterns like 's-3' or 's 3'
        m_sem = re.search(r"\bs[\s_-]*([0-9]{1,2})", s)
    try:
        semester = int(m_sem.group(1)) if m_sem else 1
    except Exception:
        semester = 1

    return program_name, semester


def get_or_create_subject_type(code: str) -> SubjectType:
    code_norm = (code or "MAJOR").strip().upper()
    st = db.session.execute(select(SubjectType).filter_by(type_code=code_norm)).scalars().first()
    if not st:
        st = SubjectType(type_code=code_norm, description=None)
        db.session.add(st)
        db.session.flush()  # get type_id
    return st


def upsert_subjects(program_name: str, path: str, default_semester: int, force_default_semester: bool = False, dry_run: bool = False):
    program = db.session.execute(select(Program).filter_by(program_name=program_name)).scalars().first()
    if not program:
        program = Program(program_name=program_name, program_duration_years=3)
        db.session.add(program)
        if dry_run:
            db.session.flush()
        else:
            db.session.commit()

    created = 0
    updated = 0

    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_headers(headers)

        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            data = {}
            for idx, key in colmap.items():
                data[key] = values[idx] if idx < len(values) else None

            subject_name = cell_to_str(data.get("subject_name")).strip()
            subject_code = cell_to_str(data.get("subject_code")).strip()
            paper_code = cell_to_str(data.get("paper_code")).strip()
            subject_type_code = cell_to_str(data.get("subject_type")).strip().upper() or "MAJOR"
            sem_val = cell_to_str(data.get("semester")).strip()
            semester = default_semester
            if not force_default_semester:
                if sem_val.isdigit():
                    try:
                        semester = int(sem_val)
                    except Exception:
                        semester = default_semester

            if not subject_name:
                # Skip blank rows
                continue

            # Ensure Subject Type exists
            st = get_or_create_subject_type(subject_type_code)

            # Upsert Subject: prefer program + subject_code; fallback to name + semester
            subj = None
            if subject_code:
                subj = db.session.execute(
                    select(Subject).filter_by(
                        program_id_fk=program.program_id,
                        subject_code=subject_code,
                    )
                ).scalars().first()
            if not subj:
                subj = db.session.execute(
                    select(Subject).filter_by(
                        program_id_fk=program.program_id,
                        subject_name=subject_name,
                        semester=semester,
                    )
                ).scalars().first()
            if not subj:
                subj = Subject(
                    program_id_fk=program.program_id,
                    subject_type_id_fk=st.type_id,
                    subject_name=subject_name,
                    subject_code=(subject_code or None),
                    paper_code=(paper_code or None),
                    semester=semester,
                )
                db.session.add(subj)
                db.session.flush()  # get subject_id
                created += 1
            else:
                subj.subject_type_id_fk = st.type_id
                if subject_code:
                    subj.subject_code = subject_code
                if paper_code:
                    subj.paper_code = paper_code
                if force_default_semester:
                    subj.semester = default_semester
                updated += 1

            # Credits
            def to_int_safe(x: Any) -> int:
                s = cell_to_str(x).strip()
                if not s:
                    return 0
                try:
                    return int(float(s))
                except Exception:
                    return 0

            th = to_int_safe(data.get("theory_credits"))
            pr = to_int_safe(data.get("practical_credits"))
            total = to_int_safe(data.get("total_credits"))
            if not total:
                total = th + pr

            cs = db.session.execute(
                select(CreditStructure).filter_by(subject_id_fk=subj.subject_id)
            ).scalars().first()
            if not cs:
                cs = CreditStructure(subject_id_fk=subj.subject_id, theory_credits=th, practical_credits=pr, total_credits=total)
                db.session.add(cs)
            else:
                cs.theory_credits = th
                cs.practical_credits = pr
                cs.total_credits = total

    except InvalidFileException:
        # Legacy .xls fallback
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        colmap = normalize_headers(headers)

        for r in range(1, sheet.nrows):
            data = {}
            for c in range(sheet.ncols):
                key = colmap.get(c)
                if not key:
                    continue
                data[key] = sheet.cell_value(r, c)

            subject_name = cell_to_str(data.get("subject_name")).strip()
            subject_code = cell_to_str(data.get("subject_code")).strip()
            paper_code = cell_to_str(data.get("paper_code")).strip()
            subject_type_code = cell_to_str(data.get("subject_type")).strip().upper() or "MAJOR"
            sem_val = cell_to_str(data.get("semester")).strip()
            semester = default_semester
            if not force_default_semester:
                if sem_val.isdigit():
                    try:
                        semester = int(sem_val)
                    except Exception:
                        semester = default_semester

            if not subject_name:
                continue

            st = get_or_create_subject_type(subject_type_code)
            subj = None
            if subject_code:
                subj = db.session.execute(
                    select(Subject).filter_by(
                        program_id_fk=program.program_id,
                        subject_code=subject_code,
                    )
                ).scalars().first()
            if not subj:
                subj = db.session.execute(
                    select(Subject).filter_by(
                        program_id_fk=program.program_id,
                        subject_name=subject_name,
                        semester=semester,
                    )
                ).scalars().first()
            if not subj:
                subj = Subject(
                    program_id_fk=program.program_id,
                    subject_type_id_fk=st.type_id,
                    subject_name=subject_name,
                    subject_code=(subject_code or None),
                    paper_code=(paper_code or None),
                    semester=semester,
                )
                db.session.add(subj)
                db.session.flush()
                created += 1
            else:
                subj.subject_type_id_fk = st.type_id
                if subject_code:
                    subj.subject_code = subject_code
                if paper_code:
                    subj.paper_code = paper_code
                if force_default_semester:
                    subj.semester = default_semester
                updated += 1

            def to_int_safe(x: Any) -> int:
                s = cell_to_str(x).strip()
                if not s:
                    return 0
                try:
                    return int(float(s))
                except Exception:
                    return 0

            th = to_int_safe(data.get("theory_credits"))
            pr = to_int_safe(data.get("practical_credits"))
            total = to_int_safe(data.get("total_credits")) or (th + pr)
            cs = db.session.execute(
                select(CreditStructure).filter_by(subject_id_fk=subj.subject_id)
            ).scalars().first()
            if not cs:
                cs = CreditStructure(subject_id_fk=subj.subject_id, theory_credits=th, practical_credits=pr, total_credits=total)
                db.session.add(cs)
            else:
                cs.theory_credits = th
                cs.practical_credits = pr
                cs.total_credits = total

    if dry_run:
        db.session.rollback()
    else:
        db.session.commit()
    print(f"Imported subjects from {path}: created={created}, updated={updated}")
    return created, updated, program.program_id


def main():
    app = create_app()
    with app.app_context():
        args = sys.argv[1:]
        if not args:
            # Default to provided BCA Sem 1 list if no args
            args = [r"c:\project\CMSv5\BCA-Sem-1-Subject-list-with-code-1.xlsx"]
        for p in args:
            program_name, semester = detect_program_and_semester(p)
            print(f"Processing: program={program_name}, default_semester={semester}, path={p}")
            upsert_subjects(program_name, p, semester)


if __name__ == "__main__":
    main()
