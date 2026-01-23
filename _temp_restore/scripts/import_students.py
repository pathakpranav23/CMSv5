import re
import sys
import os
import csv
from datetime import datetime
from typing import Dict, List
from sqlalchemy import select

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd

# Ensure project root is on sys.path when running from scripts/
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from cms_app import create_app, db
from cms_app.models import Program, Division, Student, User, ProgramDivisionPlan


HEADER_MAP: Dict[str, List[str]] = {
    # IDs
    "enrollment_no": [
        "enrollment no",
        "enrolment no",
        "enrollment number",
        "enrolment",
        "enrollment",
        "enrollment #",
        "enrollment no.",
        "roll no",
        "rollno",
        "roll number",
        "sr",
        "sr.",
        "seat no",
        "prn",
        "student id",
        "enrollment id",
    ],
    # Names (simplified to match frontend columns)
    "surname": ["surname", "last name", "family name"],
    "student_name": ["student name", "full name", "given name"],
    "father_name": ["father name", "father's name", "father’s name"],
    # Division & Semester
    "division_code": ["division", "div", "class", "section", "division name", "divison"],
    "current_semester": ["semester", "sem", "semester no", "sem no", "current sem"],
    # Contact & DOB
    "mobile": [
        "mobile",
        "mobile no",
        "mobile number",
        "phone",
        "phone no",
        "contact",
        "contact no",
        "student mobile",
        "student phone",
    ],
    "date_of_birth": ["dob", "d.o.b", "date of birth", "birthdate", "birth date"],
    # New fields
    "gender": ["gender", "sex"],
    "photo_url": ["photo", "student photo", "photo url", "image", "profile photo"],
    "permanent_address": ["permanent address", "address"],
    # Instruction medium (optional; primarily for BCom)
    "medium_tag": [
        "medium",
        "medium tag",
        "instruction medium",
        "language",
        "teach medium",
    ],
}


def load_program_mediums() -> Dict[str, Dict[str, List[str]]]:
    result: Dict[str, Dict[str, List[str]]] = {}
    try:
        p = os.path.join(PROJECT_ROOT, "DATA FOR IMPORT EXPORT", "programs.csv")
        with open(p, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("program_name") or "").strip()
                mediums = (row.get("mediums") or "").strip()
                default_medium = (row.get("default_medium") or "").strip()
                ms = [m.strip() for m in mediums.split("|") if m.strip()]
                result[name] = {"mediums": ms, "default": [default_medium]}
    except Exception:
        pass
    return result


def find_semester_from_filename(path: str) -> int:
    m = re.search(r"sem\s*(\d+)", path, flags=re.IGNORECASE)
    return int(m.group(1)) if m else 0


def detect_program_from_filename(path: str) -> str:
    s = (path or "").lower()
    # Normalize separators and punctuation for easier matching
    s = s.replace("_", " ").replace("-", " ")
    # BCA/BBA direct detection
    if "bca" in s:
        return "BCA"
    if "bba" in s:
        return "BBA"
    # BCom variants
    if ("bcom" in s) or ("b.com" in s) or ("b com" in s):
        return "B.Com"
    # Fallback
    return "BCA"


def normalize_headers(headers: List[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        h_low = (h or "").strip().lower()
        h_low = h_low.replace("’", "'").replace("—", "-").replace("–", "-")
        for key, synonyms in HEADER_MAP.items():
            # use exact match with lowercase synonyms to avoid collisions like 'name' in "father's name"
            synonyms_low = [s.lower() for s in synonyms]
            if h_low in synonyms_low:
                mapping[idx] = key
                break
    return mapping


def cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        # Excel often stores numeric IDs as float; drop .0 when applicable
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def to_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        try:
            return int(val)
        except Exception:
            return None
    try:
        s = str(val).strip()
        m = re.search(r"(\d+)", s)
        return int(m.group(1)) if m else None
    except Exception:
        return None


def import_excel(path: str, program_name: str = None, semester_hint: int = None, dry_run: bool = False):
    # Determine semester
    semester = semester_hint or find_semester_from_filename(path) or 0

    # Determine program name from filename if not provided
    if not program_name:
        program_name = detect_program_from_filename(path)

    # Ensure program exists
    program = db.session.execute(select(Program).filter_by(program_name=program_name)).scalars().first()
    if not program:
        program = Program(program_name=program_name, program_duration_years=3)
        db.session.add(program)
        db.session.commit()

    created = 0
    updated = 0
    skipped = 0
    divisions_created = 0
    errors: List[str] = []

    cfg = load_program_mediums()
    try:
        # Primary path: openpyxl for .xlsx/.xlsm
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_headers(headers)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row]
            data = {}
            for idx, key in colmap.items():
                data[key] = values[idx]

            enrollment_no = cell_to_str(data.get("enrollment_no"))
            if not enrollment_no:
                # skip rows without enrollment number
                skipped += 1
                errors.append(f"Row {row_idx}: missing enrollment_no; skipped")
                continue

            # Division (respect per-program planning)
            division_code = cell_to_str(data.get("division_code")) or "A"
            division = db.session.execute(select(Division).filter_by(program_id_fk=program.program_id, semester=semester, division_code=division_code)).scalars().first()
            if not division:
                # Determine capacity from ProgramDivisionPlan; fallback to BCA=67 else Division default
                plan = db.session.execute(select(ProgramDivisionPlan).filter_by(program_id_fk=program.program_id, semester=semester)).scalars().first()
                cap = None
                if plan:
                    try:
                        cap = int(plan.capacity_per_division)
                    except Exception:
                        cap = None
                if cap is None:
                    cap = 67 if (program.program_name or "").upper() == "BCA" else (Division.capacity.default.arg if hasattr(Division.capacity, 'default') else 60)
                division = Division(program_id_fk=program.program_id, semester=semester, division_code=division_code, capacity=cap)
                db.session.add(division)
                if not dry_run:
                    db.session.commit()
                divisions_created += 1
            else:
                # Align capacity with planning when available; avoid uniform forcing
                plan = db.session.execute(select(ProgramDivisionPlan).filter_by(program_id_fk=program.program_id, semester=semester)).scalars().first()
                if plan:
                    try:
                        cap = int(plan.capacity_per_division)
                        if division.capacity != cap:
                            division.capacity = cap
                            if not dry_run:
                                db.session.commit()
                    except Exception:
                        pass

            # Student fields (use only provided columns; no composition)
            surname = cell_to_str(data.get("surname"))
            student_name = cell_to_str(data.get("student_name"))
            mobile = cell_to_str(data.get("mobile"))
            father_name = cell_to_str(data.get("father_name"))
            gender = cell_to_str(data.get("gender")).capitalize()
            if gender not in ("Male", "Female", "Other", ""):
                gender = ""
            photo_url = cell_to_str(data.get("photo_url"))
            permanent_address = cell_to_str(data.get("permanent_address"))
            # Optional medium parsing
            medium_raw = cell_to_str(data.get("medium_tag")).strip().lower()
            medium_map = {
                "": "",
                "general": "General",
                "eng": "English",
                "english": "English",
                "e": "English",
                "guj": "Gujarati",
                "gujarati": "Gujarati",
                "g": "Gujarati",
            }
            medium_tag = medium_map.get(medium_raw, "")
            allowed = []
            default_m = ""
            try:
                cfg_row = cfg.get(program.program_name or "") or {}
                allowed = cfg_row.get("mediums") or []
                default_list = cfg_row.get("default") or []
                default_m = default_list[0] if default_list else ""
            except Exception:
                allowed = []
                default_m = ""
            if not medium_tag:
                medium_tag = default_m
            if allowed and medium_tag and medium_tag not in allowed:
                errors.append(f"Row {row_idx}: medium '{medium_tag}' not allowed for {program.program_name}")
                medium_tag = default_m

            dob_val = data.get("date_of_birth")
            dob = None
            if isinstance(dob_val, datetime):
                dob = dob_val.date()
            elif isinstance(dob_val, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        dob = datetime.strptime(dob_val.strip(), fmt).date()
                        break
                    except Exception:
                        pass

            current_semester = semester or to_int(data.get("current_semester"))

            student = db.session.execute(select(Student).filter_by(enrollment_no=enrollment_no)).scalars().first()
            if not student:
                student = Student(
                    enrollment_no=enrollment_no,
                    program_id_fk=program.program_id,
                    division_id_fk=division.division_id,
                    surname=surname,
                    student_name=student_name,
                    father_name=father_name,
                    mobile=mobile,
                    date_of_birth=dob,
                    gender=gender,
                    medium_tag=None,
                    photo_url=photo_url,
                    permanent_address=permanent_address,
                    current_semester=current_semester,
                )
                db.session.add(student)
                created += 1
            else:
                student.program_id_fk = program.program_id
                student.division_id_fk = division.division_id
                student.surname = surname or student.surname
                student.student_name = student_name or student.student_name
                student.mobile = mobile or student.mobile
                student.father_name = father_name or student.father_name
                student.date_of_birth = dob or student.date_of_birth
                student.gender = gender or student.gender
                student.photo_url = photo_url or student.photo_url
                student.permanent_address = permanent_address or student.permanent_address
                student.current_semester = current_semester or student.current_semester
                updated += 1
            # Assign medium with BCom defaulting to General when absent
            try:
                student.medium_tag = medium_tag or (student.medium_tag or None)
            except Exception:
                student.medium_tag = medium_tag or (student.medium_tag or None)
                errors.append(f"Row {row_idx}: failed to compute medium_tag due to data format")

    except InvalidFileException:
        # Fallback path: xlrd for legacy .xls or malformed files
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        colmap = normalize_headers(headers)

        for r in range(1, sheet.nrows):
            data = {}
            for c in range(sheet.ncols):
                key = colmap.get(c)
                if not key:
                    continue
                val = sheet.cell_value(r, c)
                # Convert Excel serial date
                if key == "date_of_birth" and sheet.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, book.datemode).date()
                    except Exception:
                        pass
                data[key] = val

            enrollment_no = cell_to_str(data.get("enrollment_no"))
            if not enrollment_no:
                skipped += 1
                errors.append(f"Row {r+1}: missing enrollment_no; skipped")
                continue

            division_code = cell_to_str(data.get("division_code")) or "A"
            division = db.session.execute(select(Division).filter_by(program_id_fk=program.program_id, semester=semester, division_code=division_code)).scalars().first()
            if not division:
                plan = db.session.execute(select(ProgramDivisionPlan).filter_by(program_id_fk=program.program_id, semester=semester)).scalars().first()
                cap = None
                if plan:
                    try:
                        cap = int(plan.capacity_per_division)
                    except Exception:
                        cap = None
                if cap is None:
                    cap = 67 if (program.program_name or "").upper() == "BCA" else (Division.capacity.default.arg if hasattr(Division.capacity, 'default') else 60)
                division = Division(program_id_fk=program.program_id, semester=semester, division_code=division_code, capacity=cap)
                db.session.add(division)
                if not dry_run:
                    db.session.commit()
                divisions_created += 1

            surname = cell_to_str(data.get("surname"))
            student_name = cell_to_str(data.get("student_name"))
            mobile = cell_to_str(data.get("mobile"))
            father_name = cell_to_str(data.get("father_name"))
            gender = cell_to_str(data.get("gender")).capitalize()
            if gender not in ("Male", "Female", "Other", ""):
                gender = ""
            photo_url = cell_to_str(data.get("photo_url"))
            permanent_address = cell_to_str(data.get("permanent_address"))
            # Optional medium parsing
            medium_raw = cell_to_str(data.get("medium_tag")).strip().lower()
            medium_map = {
                "": "",
                "general": "General",
                "eng": "English",
                "english": "English",
                "e": "English",
                "guj": "Gujarati",
                "gujarati": "Gujarati",
                "g": "Gujarati",
            }
            medium_tag = medium_map.get(medium_raw, "")
            allowed = []
            default_m = ""
            try:
                cfg_row = cfg.get(program.program_name or "") or {}
                allowed = cfg_row.get("mediums") or []
                default_list = cfg_row.get("default") or []
                default_m = default_list[0] if default_list else ""
            except Exception:
                allowed = []
                default_m = ""
            if not medium_tag:
                medium_tag = default_m
            if allowed and medium_tag and medium_tag not in allowed:
                errors.append(f"Row {r+1}: medium '{medium_tag}' not allowed for {program.program_name}")
                medium_tag = default_m

            dob_val = data.get("date_of_birth")
            dob = None
            if isinstance(dob_val, datetime):
                dob = dob_val
            elif isinstance(dob_val, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        dob = datetime.strptime(dob_val.strip(), fmt).date()
                        break
                    except Exception:
                        pass

            current_semester = semester or to_int(data.get("current_semester"))

            student = db.session.execute(select(Student).filter_by(enrollment_no=enrollment_no)).scalars().first()
            if not student:
                student = Student(
                    enrollment_no=enrollment_no,
                    program_id_fk=program.program_id,
                    division_id_fk=division.division_id,
                    surname=surname,
                    student_name=student_name,
                    father_name=father_name,
                    mobile=mobile,
                    date_of_birth=dob,
                    gender=gender,
                    medium_tag=None,
                    photo_url=photo_url,
                    permanent_address=permanent_address,
                    current_semester=current_semester,
                )
                db.session.add(student)
                created += 1
            else:
                student.program_id_fk = program.program_id
                student.division_id_fk = division.division_id
                student.surname = surname or student.surname
                student.student_name = student_name or student.student_name
                student.mobile = mobile or student.mobile
                student.father_name = father_name or student.father_name
                student.date_of_birth = dob or student.date_of_birth
                student.gender = gender or student.gender
                student.photo_url = photo_url or student.photo_url
                student.permanent_address = permanent_address or student.permanent_address
                student.current_semester = current_semester or student.current_semester
                updated += 1
            # Assign medium with BCom defaulting to General when absent
            try:
                student.medium_tag = medium_tag or (student.medium_tag or None)
            except Exception:
                student.medium_tag = medium_tag or (student.medium_tag or None)
                errors.append(f"Row {r+1}: failed to compute medium_tag due to data format")

    if not dry_run:
        db.session.commit()
    else:
        db.session.rollback()
    print(f"Imported from {path}: created={created}, updated={updated}, skipped={skipped}, divisions_created={divisions_created}, errors={len(errors)}")
    # Return a detailed report for UI display
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors_count": len(errors),
        "errors": errors,
        "divisions_created": divisions_created,
        "program_name": program.program_name,
        "program_id": program.program_id,
        "semester": semester,
        "path": path,
    }


def main():
    app = create_app()
    with app.app_context():
        args = sys.argv[1:]
        if not args:
            # Default to BCom bulk import files when no args provided
            args = [
                r"c:\project\CMSv5\B.Com\Bulk BCom Student Import for Semester 3.xlsx",
                r"c:\project\CMSv5\B.Com\Bulk BCom Student Import for Semester 5.xlsx",
            ]
        for p in args:
            # Detect program from filename and use semester hints from filenames
            program_name = detect_program_from_filename(p)
            sem_hint = find_semester_from_filename(p)
            import_excel(p, program_name=program_name, semester_hint=sem_hint or None)


if __name__ == "__main__":
    main()