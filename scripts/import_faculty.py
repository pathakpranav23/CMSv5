import re
import json
import sys
import os
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd

# Ensure project root is on sys.path when running from scripts/
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from cms_app import create_app, db
from cms_app.models import Program, User, Faculty
from sqlalchemy import select
from werkzeug.security import generate_password_hash


HEADER_MAP: Dict[str, List[str]] = {
    "full_name": [
        "name",
        "full name",
        "faculty name",
        "teacher name",
    ],
    "email": [
        "email",
        "email id",
        "e-mail",
        "mail",
        "username",
    ],
    "mobile": [
        "mobile",
        "mobile no",
        "phone",
        "contact",
        "contact no",
    ],
    "designation": [
        "designation",
        "title",
        "role",
        "position",
    ],
    "department": [
        "department",
        "dept",
        "programme",
        "program",
        "program name",
    ],
}


def normalize_headers(headers: List[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        h_low = (h or "").strip().lower()
        h_low = h_low.replace("’", "'").replace("—", "-").replace("–", "-")
        for key, synonyms in HEADER_MAP.items():
            synonyms_low = [s.lower() for s in synonyms]
            if h_low in synonyms_low:
                mapping[idx] = key
                break
    return mapping


def cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def derive_name_from_email(email: str) -> str:
    if not email:
        return ""
    try:
        local = email.split("@")[0]
        # drop digits and normalize separators to spaces
        local = re.sub(r"\d+", "", local)
        local = re.sub(r"[._-]+", " ", local)
        local = local.strip()
        if not local:
            return ""
        parts = [p for p in local.split(" ") if p]
        return " ".join(s.capitalize() for s in parts)
    except Exception:
        return ""


def detect_program_from_filename(path: str) -> str:
    s = (path or "").lower()
    s = s.replace("_", " ").replace("-", " ")
    if "bca" in s:
        return "BCA"
    if "bba" in s:
        return "BBA"
    if ("bcom" in s) or ("b.com" in s) or ("b com" in s):
        if ("eng" in s) or ("english" in s):
            return "BCom (English)"
        if ("guj" in s) or ("gujarati" in s):
            return "BCom (Gujarati)"
        return "BCOM"
    return "BCA"


def upsert_faculty(program_name: str, path: str):
    program = db.session.execute(
        select(Program).filter_by(program_name=program_name)
    ).scalars().first()
    if not program:
        program = Program(program_name=program_name, program_duration_years=3)
        db.session.add(program)
        db.session.commit()

    created = 0
    updated = 0

    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_headers(headers)

        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            data = {}
            for idx, key in colmap.items():
                data[key] = values[idx]

            # Capture all original columns keyed by their header
            raw_map = {}
            for i, h in enumerate(headers):
                raw_map[str(h or "")] = cell_to_str(values[i] if i < len(values) else "")

            full_name = cell_to_str(data.get("full_name"))
            email = cell_to_str(data.get("email")).lower()
            mobile = cell_to_str(data.get("mobile"))
            designation = cell_to_str(data.get("designation"))
            department = cell_to_str(data.get("department")) or program_name

            if not full_name and not email:
                # skip blank rows
                continue

            # Choose username: prefer email, else mobile, else name
            username = email or (mobile if mobile else full_name)
            username = username.strip()
            if not username:
                username = full_name or "faculty"

            user = db.session.execute(
                select(User).filter_by(username=username)
            ).scalars().first()
            if not user:
                # Default password for imported faculty: Password123
                # Must change password on first login
                user = User(
                    username=username, 
                    role="Faculty", 
                    program_id_fk=program.program_id,
                    password_hash=generate_password_hash("Password123"),
                    must_change_password=True
                )
                db.session.add(user)
                db.session.flush()  # get user_id
            else:
                user.role = "Faculty"
                user.program_id_fk = program.program_id
                # Ensure existing users also have to change password if they haven't logged in? 
                # No, only force for new users or if explicitly reset.


            # Upsert Faculty record using email if present, else username
            fac_q = None
            if email:
                fac_q = db.session.execute(
                    select(Faculty).filter_by(email=email)
                ).scalars().first()
            if not fac_q:
                fac_q = db.session.execute(
                    select(Faculty).filter_by(user_id_fk=user.user_id)
                ).scalars().first()

            if not fac_q:
                fac = Faculty(
                    user_id_fk=user.user_id,
                    program_id_fk=program.program_id,
                    full_name=(full_name if full_name and full_name != email else (derive_name_from_email(email) or username)),
                    email=email or None,
                    mobile=mobile or None,
                    designation=designation or None,
                    department=department or None,
                    extra_data=json.dumps(raw_map, ensure_ascii=False),
                )
                db.session.add(fac)
                created += 1
            else:
                fac_q.program_id_fk = program.program_id
                # Update name if provided or if existing equals the email
                if full_name and full_name != email:
                    fac_q.full_name = full_name
                elif not fac_q.full_name or fac_q.full_name == fac_q.email:
                    derived = derive_name_from_email(email)
                    if derived:
                        fac_q.full_name = derived
                fac_q.email = email or fac_q.email
                fac_q.mobile = mobile or fac_q.mobile
                fac_q.designation = designation or fac_q.designation
                fac_q.department = department or fac_q.department
                fac_q.extra_data = json.dumps(raw_map, ensure_ascii=False)
                updated += 1

    except InvalidFileException:
        # Fallback to xlrd for legacy .xls
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        colmap = normalize_headers(headers)

        for r in range(1, sheet.nrows):
            data = {}
            for c in range(sheet.ncols):
                key = colmap.get(c)
                if not key:
                    continue
                val = sheet.cell_value(r, c)
                data[key] = val

            # Capture all original columns keyed by their header
            raw_map = {}
            for c in range(sheet.ncols):
                h = headers[c] if c < len(headers) else ""
                raw_map[str(h or "")] = cell_to_str(sheet.cell_value(r, c))

            full_name = cell_to_str(data.get("full_name"))
            email = cell_to_str(data.get("email")).lower()
            mobile = cell_to_str(data.get("mobile"))
            designation = cell_to_str(data.get("designation"))
            department = cell_to_str(data.get("department")) or program_name

            if not full_name and not email:
                continue

            username = email or (mobile if mobile else full_name)
            username = username.strip() or (full_name or "faculty")

            user = db.session.execute(
                select(User).filter_by(username=username)
            ).scalars().first()
            if not user:
                user = User(username=username, role="Faculty", program_id_fk=program.program_id)
                db.session.add(user)
                db.session.flush()
            else:
                user.role = "Faculty"
                user.program_id_fk = program.program_id

            fac_q = None
            if email:
                fac_q = db.session.execute(
                    select(Faculty).filter_by(email=email)
                ).scalars().first()
            if not fac_q:
                fac_q = db.session.execute(
                    select(Faculty).filter_by(user_id_fk=user.user_id)
                ).scalars().first()

            if not fac_q:
                fac = Faculty(
                    user_id_fk=user.user_id,
                    program_id_fk=program.program_id,
                    full_name=(full_name if full_name and full_name != email else (derive_name_from_email(email) or username)),
                    email=email or None,
                    mobile=mobile or None,
                    designation=designation or None,
                    department=department or None,
                    extra_data=json.dumps(raw_map, ensure_ascii=False),
                )
                db.session.add(fac)
                created += 1
            else:
                fac_q.program_id_fk = program.program_id
                if full_name and full_name != email:
                    fac_q.full_name = full_name
                elif not fac_q.full_name or fac_q.full_name == fac_q.email:
                    derived = derive_name_from_email(email)
                    if derived:
                        fac_q.full_name = derived
                fac_q.email = email or fac_q.email
                fac_q.mobile = mobile or fac_q.mobile
                fac_q.designation = designation or fac_q.designation
                fac_q.department = department or fac_q.department
                fac_q.extra_data = json.dumps(raw_map, ensure_ascii=False)
                updated += 1

    db.session.commit()
    print(f"Imported faculty from {path}: created={created}, updated={updated}")


def main():
    app = create_app()
    with app.app_context():
        args = sys.argv[1:]
        if not args:
            # Default to provided path
            args = [r"c:\project\CMSv5\BCA Staff info.xlsx"]
        for p in args:
            # Detect program from filename with BCom English/Gujarati support
            program_name = detect_program_from_filename(p)
            upsert_faculty(program_name, p)


if __name__ == "__main__":
    main()
