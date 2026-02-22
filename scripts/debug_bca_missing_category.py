import os
import sys

from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from cms_app import create_app, db  # noqa: E402
from cms_app.models import Program, Student  # noqa: E402
from sqlalchemy import select, func  # noqa: E402


def cell_to_str(val):
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def main():
    app = create_app()
    with app.app_context():
        program = db.session.execute(
            select(Program).filter(func.lower(Program.program_name) == "bca")
        ).scalars().first()
        if not program:
            print("No BCA program found")
            return

        q = (
            select(Student)
            .filter(Student.program_id_fk == program.program_id)
            .filter(
                (Student.category == None)
                | (func.trim(Student.category) == "")
            )
            .order_by(Student.current_semester.asc(), Student.enrollment_no.asc())
        )
        students = db.session.execute(q).scalars().all()
        print(f"BCA students with missing/blank category: {len(students)}")
        for s in students:
            name = f"{getattr(s, 'student_name', '')} {getattr(s, 'surname', '')}".strip()
            print(
                f"Sem {getattr(s, 'current_semester', None)} | "
                f"{getattr(s, 'enrollment_no', '')} | "
                f"{name} | medium={getattr(s, 'medium_tag', '')}"
            )

        excel_path = r"c:\project\CMSv5\DATA FOR IMPORT EXPORT\BCA\BCA Sem 4 Bulk Student Data 2026 22 feb.xlsx"
        if not os.path.exists(excel_path):
            print("Excel file not found:", excel_path)
            return

        wb = load_workbook(excel_path, data_only=True)
        mapping = {}
        for ws in wb.worksheets:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            print("SHEET:", ws.title)
            for i, h in enumerate(header_row):
                print(f"  COL {i}: {h}")
            enr_idx = None
            cat_idx = None
            for idx, val in enumerate(header_row):
                if not val:
                    continue
                h = str(val).strip().lower()
                if enr_idx is None and (
                    "enrollment" in h
                    or "enrolment" in h
                    or "seat no" in h
                    or "roll no" in h
                    or "prn" in h
                ):
                    enr_idx = idx
                if cat_idx is None and (
                    "category" in h
                    or "caste" in h
                    or "reservation" in h
                ):
                    cat_idx = idx
            print("  enr_idx:", enr_idx, "cat_idx:", cat_idx)
            if enr_idx is None or cat_idx is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if enr_idx >= len(row):
                    continue
                enr = cell_to_str(row[enr_idx])
                if not enr:
                    continue
                cat_val = ""
                if cat_idx < len(row):
                    cat_val = cell_to_str(row[cat_idx])
                mapping[enr] = cat_val

        print()
        print("Excel categories (only where found) for BCA students with missing category in DB:")
        found_any = False
        for s in students:
            enr = cell_to_str(getattr(s, "enrollment_no", ""))
            name = f"{getattr(s, 'student_name', '')} {getattr(s, 'surname', '')}".strip()
            excel_cat = mapping.get(enr, "")
            if excel_cat:
                found_any = True
                print(
                    f"{enr} | {name} | sem={getattr(s, 'current_semester', None)} | excel_category={excel_cat}"
                )
        if not found_any:
            print("No matching categories found in this Excel file for the missing-category BCA students.")


if __name__ == "__main__":
    main()
