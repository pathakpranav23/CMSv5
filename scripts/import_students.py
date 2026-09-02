import re
import sys
import os
import csv
from datetime import datetime
from typing import Dict, List
from sqlalchemy import func, select
from werkzeug.security import generate_password_hash
from sqlalchemy.exc import OperationalError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd

# Ensure project root is on sys.path when running from scripts/
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from cms_app import create_app, db
from cms_app.models import Program, Division, Student, User, ProgramDivisionPlan, ProgramIntakeBatch, Institute


HEADER_MAP: Dict[str, List[str]] = {
    # IDs
    "enrollment_no": [
        "enrollment no",
        "enrolment no",
        "enrollment number",
        "enrolment",
        "enrollment",
        "enrollment #",
        "enrollment no.",
        "sr",
        "sr.",
        "seat no",
        "prn",
        "student id",
        "enrollment id",
    ],
    "roll_no": [
        "roll no",
        "rollno",
        "roll number",
    ],
    # Names (simplified to match frontend columns)
    "last_name": ["surname", "last name", "family name"],
    "first_name": ["student name", "full name", "given name"],
    "father_name": ["father name", "father's name", "father’s name"],
    # Division & Semester
    "division_code": ["division", "div", "class", "section", "division name", "divison"],
    "current_semester": ["semester", "sem", "semester no", "sem no", "current sem"],
    "admission_academic_year": ["admission academic year", "admission year", "batch year", "admission batch"],
    # Contact & DOB
    "mobile": [
        "mobile",
        "mobile no",
        "mobile number",
        "phone",
        "phone no",
        "contact",
        "contact no",
        "student mobile",
        "student phone",
    ],
    "date_of_birth": ["dob", "d.o.b", "date of birth", "birthdate", "birth date"],
    # New fields
    "gender": ["gender", "sex"],
    "photo_url": ["photo", "student photo", "photo url", "image", "profile photo"],
    "permanent_address": ["permanent address", "address"],
    "aadhar_no": [
        "aadhar card number",
        "aadhaar card number",
        "aadhar no",
        "aadhaar no",
        "aadhar",
        "aadhaar",
    ],
    "category": [
        "category",
        "caste category",
        "caste",
        "reservation category",
    ],
    # Instruction medium (optional; primarily for BCom)
    "medium_tag": [
        "medium",
        "medium tag",
        "instruction medium",
        "language",
        "teach medium",
    ],
}


def load_program_mediums() -> Dict[str, Dict[str, List[str]]]:
    result: Dict[str, Dict[str, List[str]]] = {}
    try:
        p = os.path.join(PROJECT_ROOT, "DATA FOR IMPORT EXPORT", "programs.csv")
        with open(p, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("program_name") or "").strip()
                mediums = (row.get("mediums") or "").strip()
                default_medium = (row.get("default_medium") or "").strip()
                ms = [m.strip() for m in mediums.split("|") if m.strip()]
                result[name] = {"mediums": ms, "default": [default_medium]}
    except Exception:
        pass
    return result


def find_semester_from_filename(path: str) -> int:
    """
    Tries to guess semester from filename like 'BCA Sem 1.xlsx' or 'Semester 4'.
    Returns 0 if not found.
    """
    # Look for 'sem' or 'semester' followed optionally by space, then digits
    m = re.search(r"(?:sem|semester)\s*(\d+)", path, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    return 0


def detect_program_from_filename(path: str) -> str:
    s = (path or "").lower()
    # Normalize separators and punctuation for easier matching
    s = s.replace("_", " ").replace("-", " ")
    # BCA/BBA direct detection
    if "bca" in s:
        return "BCA"
    if "bba" in s:
        return "BBA"
    # BCom variants
    if ("bcom" in s) or ("b.com" in s) or ("b com" in s):
        return "B.Com"
    # Fallback
    return "BCA"


def normalize_headers(headers: List[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        h_low = (h or "").strip().lower()
        h_low = h_low.replace("_", " ").replace("’", "'").replace("—", "-").replace("–", "-")
        for key, synonyms in HEADER_MAP.items():
            # use exact match with lowercase synonyms to avoid collisions like 'name' in "father's name"
            synonyms_low = [s.lower() for s in synonyms] + [key.replace("_", " ").lower()]
            if h_low in synonyms_low:
                mapping[idx] = key
                break
    return mapping


def cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        # Excel often stores numeric IDs as float; drop .0 when applicable
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def normalize_academic_year(val) -> str:
    """Return an academic year in YYYY-YY form when the value is unambiguous.

    College spreadsheets commonly store only the admission start year (for
    example, 2026).  Treat that as the 2026-27 academic year while preserving
    other values so genuine conflicts are still rejected by the importer.
    """
    value = cell_to_str(val).strip()
    if re.fullmatch(r"\d{4}", value):
        start_year = int(value)
        return f"{start_year}-{(start_year + 1) % 100:02d}"
    match = re.fullmatch(r"(\d{4})\s*[-/]\s*(\d{2}|\d{4})", value)
    if match:
        start_year = int(match.group(1))
        end_year = int(match.group(2))
        if len(match.group(2)) == 4:
            end_year %= 100
        if end_year == (start_year + 1) % 100:
            return f"{start_year}-{end_year:02d}"
    return value


def to_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        try:
            return int(val)
        except Exception:
            return None
    try:
        s = str(val).strip()
        m = re.search(r"(\d+)", s)
        return int(m.group(1)) if m else None
    except Exception:
        return None


def ensure_student_user(enrollment_no, mobile, program_id, trust_id=None):
    """
    Ensures a User account exists for the student.
    Strategy:
      - Username: Mobile (if >= 10 digits) else Enrollment No
      - Password: Mobile (if >= 10 digits) else Enrollment No
      - Force Password Change: True
    Returns: user_id
    """
    mobile_digits = "".join(ch for ch in str(mobile) if ch.isdigit()) if mobile else ""

    if len(mobile_digits) >= 10:
        username = mobile_digits
        password_raw = mobile_digits
    else:
        username = str(enrollment_no)
        password_raw = str(enrollment_no)

    attempts = 0
    while True:
        attempts += 1
        try:
            user = db.session.execute(select(User).filter_by(username=username)).scalars().first()
            if user:
                try:
                    if hasattr(user, "is_active") and not user.is_active:
                        user.is_active = True
                except Exception:
                    pass
            if not user:
                user = User(
                    username=username,
                    password_hash=generate_password_hash(password_raw),
                    role="student",
                    program_id_fk=program_id,
                    mobile=mobile_digits,
                    must_change_password=True,
                    is_active=True,
                    trust_id_fk=trust_id,
                )
                db.session.add(user)
                db.session.flush()
            else:
                try:
                    if hasattr(user, "program_id_fk") and program_id:
                        user.program_id_fk = program_id
                except Exception:
                    pass
                try:
                    if trust_id and hasattr(user, "trust_id_fk"):
                        user.trust_id_fk = trust_id
                except Exception:
                    pass
            return user.user_id
        except OperationalError as e:
            msg = str(e).lower()
            try:
                db.session.rollback()
            except Exception:
                pass
            if "database is locked" in msg and attempts < 6:
                import time
                time.sleep(0.2 * attempts)
                continue
            print(f"Error ensuring user for student {enrollment_no}: {e}")
            return None
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            print(f"Error ensuring user for student {enrollment_no}: {e}")
            return None



def import_excel(
    path: str,
    program_id: int = None,
    trust_id: int = None,
    program_name: str = None,
    semester_hint: int = None,
    dry_run: bool = False,
    import_mode: str = "existing",
    admission_academic_year_hint: str = None,
    medium_hint: str = None,
    division_hint: str = None,
    replace_scope: bool = False,
):
    # Determine semester
    semester = semester_hint or find_semester_from_filename(path) or 0

    program = None
    if program_id:
        program = db.session.get(Program, int(program_id))
        if not program:
            raise ValueError("Program not found.")
        if trust_id:
            ok = db.session.execute(
                select(func.count())
                .select_from(Program)
                .join(Institute, Program.institute_id_fk == Institute.institute_id)
                .filter(Program.program_id == program.program_id)
                .filter(Institute.trust_id_fk == int(trust_id))
            ).scalar() or 0
            if ok <= 0:
                raise ValueError("Program is not in the selected tenant workspace.")
    else:
        # Determine program name from filename if not provided
        if not program_name:
            program_name = detect_program_from_filename(path)

        # Ensure program exists (legacy behavior)
        program = db.session.execute(select(Program).filter_by(program_name=program_name)).scalars().first()
        if not program:
            program = Program(program_name=program_name, program_duration_years=3)
            db.session.add(program)
            db.session.flush()

    import_mode = (import_mode or "").strip().lower()
    if import_mode not in {"new", "existing"}:
        raise ValueError("Import type must be new admission or existing students.")
    admission_academic_year_hint = (admission_academic_year_hint or "").strip()
    year_match = re.fullmatch(r"(\d{4})-(\d{2})", admission_academic_year_hint)
    if not year_match or int(year_match.group(2)) != (int(year_match.group(1)) + 1) % 100:
        raise ValueError("A valid admission academic year is required, for example 2025-26.")
    if not semester or semester < 1:
        raise ValueError("An exact current semester is required.")
    maximum_semester = max(int(program.program_duration_years or 0), 1) * 2
    if semester > maximum_semester:
        raise ValueError(
            f"Semester {semester} is outside the configured {program.program_duration_years or 1}-year "
            f"structure for {program.program_name}. Correct the program duration before importing."
        )
    medium_hint = (medium_hint or "").strip()
    division_hint = (division_hint or "").strip().upper()
    selected_intake_batch = None
    if import_mode == "new":
        batches = db.session.execute(
            select(ProgramIntakeBatch).filter_by(
                program_id_fk=program.program_id,
                admission_academic_year=admission_academic_year_hint,
                status="active",
            )
        ).scalars().all()
        selected_intake_batch = next(
            (batch for batch in batches if (batch.medium_tag or "").strip().lower() == medium_hint.lower()),
            None,
        ) or next((batch for batch in batches if not (batch.medium_tag or "").strip()), None)
        if not selected_intake_batch:
            raise ValueError("New admissions require an active approved-intake record for the selected program, academic year, and MOI.")

    created = 0
    updated = 0
    skipped = 0
    deleted = 0
    divisions_created = 0
    errors: List[str] = []
    preview_rows = []
    preview_total = 0
    preview_limit = 200

    def add_preview(row_number, data, action, status="valid", reason=""):
        nonlocal preview_total
        preview_total += 1
        if len(preview_rows) >= preview_limit:
            return
        last_name = cell_to_str(data.get("last_name"))
        first_name = cell_to_str(data.get("first_name"))
        preview_rows.append({
            "row_number": row_number,
            "enrollment_no": cell_to_str(data.get("enrollment_no")),
            "roll_no": cell_to_str(data.get("roll_no")),
            "student_name": " ".join(part for part in (last_name, first_name) if part).strip() or "—",
            "admission_academic_year": normalize_academic_year(data.get("admission_academic_year")) or admission_academic_year_hint,
            "semester": to_int(data.get("current_semester")) or semester,
            "medium_tag": cell_to_str(data.get("medium_tag")) or medium_hint or "—",
            "division_code": cell_to_str(data.get("division_code")).upper() or division_hint or "Unassigned",
            "action": action,
            "status": status,
            "reason": reason,
        })

    # Get all existing students for this program and semester (for potential deletion)
    existing_q = select(Student).filter_by(program_id_fk=program.program_id, current_semester=semester)
    if trust_id:
        existing_q = existing_q.filter(Student.trust_id_fk == int(trust_id))
    existing_students = db.session.execute(existing_q).scalars().all()
    processed_enrollments = set()
    mediums_seen = set()

    cfg = load_program_mediums()
    try:
        # Primary path: openpyxl for .xlsx/.xlsm
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_headers(headers)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row]
            data = {}
            for idx, key in colmap.items():
                data[key] = values[idx]

            enrollment_no = cell_to_str(data.get("enrollment_no"))
            if not enrollment_no:
                try:
                    if not any(cell_to_str(v) for v in data.values()):
                        continue
                except Exception:
                    pass
                # skip rows without enrollment number
                skipped += 1
                errors.append(f"Row {row_idx}: missing enrollment_no; skipped")
                add_preview(row_idx, data, "skip", "error", "Missing enrollment number.")
                continue

            row_semester = to_int(data.get("current_semester"))
            if row_semester and row_semester != semester:
                skipped += 1
                errors.append(f"Row {row_idx}: semester {row_semester} conflicts with selected semester {semester}; skipped")
                add_preview(row_idx, data, "skip", "error", f"Semester {row_semester} conflicts with selected semester {semester}.")
                continue
            row_admission_year_raw = cell_to_str(data.get("admission_academic_year")).strip()
            row_admission_year = normalize_academic_year(row_admission_year_raw)
            if row_admission_year and row_admission_year != admission_academic_year_hint:
                skipped += 1
                errors.append(f"Row {row_idx}: admission year '{row_admission_year_raw}' conflicts with selected '{admission_academic_year_hint}'; skipped")
                add_preview(row_idx, data, "skip", "error", f"Admission year {row_admission_year_raw} conflicts with selected {admission_academic_year_hint}.")
                continue
            processed_enrollments.add(enrollment_no)

            # Division (respect selected import scope and per-program planning)
            row_division_code = cell_to_str(data.get("division_code")).strip().upper()
            if row_division_code and division_hint and row_division_code != division_hint:
                skipped += 1
                errors.append(f"Row {row_idx}: division '{row_division_code}' conflicts with selected '{division_hint}'; skipped")
                add_preview(row_idx, data, "skip", "error", f"Division {row_division_code} conflicts with selected {division_hint}.")
                continue
            division_code = row_division_code or division_hint
            division = db.session.execute(select(Division).filter_by(program_id_fk=program.program_id, semester=semester, division_code=division_code)).scalars().first() if division_code else None
            if division_code and not division:
                # Determine capacity from ProgramDivisionPlan; fallback to BCA=67 else Division default
                plan = db.session.execute(select(ProgramDivisionPlan).filter_by(program_id_fk=program.program_id, semester=semester)).scalars().first()
                cap = None
                if plan:
                    try:
                        cap = int(plan.capacity_per_division)
                    except Exception:
                        cap = None
                if cap is None:
                    cap = 67 if (program.program_name or "").upper() == "BCA" else (Division.capacity.default.arg if hasattr(Division.capacity, 'default') else 60)
                division = Division(program_id_fk=program.program_id, semester=semester, division_code=division_code, capacity=cap)
                db.session.add(division)
                db.session.flush()
                divisions_created += 1
            else:
                # Align capacity with planning when available; avoid uniform forcing
                plan = db.session.execute(select(ProgramDivisionPlan).filter_by(program_id_fk=program.program_id, semester=semester)).scalars().first()
                if plan:
                    try:
                        cap = int(plan.capacity_per_division)
                        if division.capacity != cap:
                            division.capacity = cap
                    except Exception:
                        pass

            # Student fields (use only provided columns; no composition)
            surname = cell_to_str(data.get("last_name"))
            student_name = cell_to_str(data.get("first_name"))
            mobile = cell_to_str(data.get("mobile"))
            father_name = cell_to_str(data.get("father_name"))
            gender = cell_to_str(data.get("gender")).capitalize()
            if gender not in ("Male", "Female", "Other", ""):
                gender = ""
            photo_url = cell_to_str(data.get("photo_url"))
            permanent_address = cell_to_str(data.get("permanent_address"))
            # Optional medium parsing
            row_medium_raw = cell_to_str(data.get("medium_tag")).strip()
            if row_medium_raw and medium_hint and row_medium_raw.lower() != medium_hint.lower():
                skipped += 1
                errors.append(f"Row {row_idx}: MOI '{row_medium_raw}' conflicts with selected '{medium_hint}'; skipped")
                add_preview(row_idx, data, "skip", "error", f"MOI {row_medium_raw} conflicts with selected {medium_hint}.")
                continue
            medium_raw = (row_medium_raw or medium_hint).lower()
            medium_map = {
                "": "",
                "general": "General",
                "eng": "English",
                "english": "English",
                "e": "English",
                "guj": "Gujarati",
                "gujarati": "Gujarati",
                "g": "Gujarati",
            }
            medium_tag = medium_map.get(medium_raw, "")
            allowed = []
            default_m = ""
            try:
                cfg_row = cfg.get(program.program_name or "") or {}
                allowed = cfg_row.get("mediums") or []
                default_list = cfg_row.get("default") or []
                default_m = default_list[0] if default_list else ""
            except Exception:
                allowed = []
                default_m = ""
            if not medium_tag:
                medium_tag = default_m
            if allowed and medium_tag and medium_tag not in allowed:
                errors.append(f"Row {row_idx}: medium '{medium_tag}' not allowed for {program.program_name}")
                medium_tag = default_m
            mediums_seen.add((medium_tag or "").strip())

            dob_val = data.get("date_of_birth")
            dob = None
            if isinstance(dob_val, datetime):
                dob = dob_val.date()
            elif isinstance(dob_val, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        dob = datetime.strptime(dob_val.strip(), fmt).date()
                        break
                    except Exception:
                        pass

            current_semester = semester
            admission_academic_year = admission_academic_year_hint
            intake_batch = selected_intake_batch if import_mode == "new" else None
            roll_no = cell_to_str(data.get("roll_no"))
            aadhar_no = cell_to_str(data.get("aadhar_no"))
            category = cell_to_str(data.get("category"))

            # Ensure User exists
            user_id = ensure_student_user(enrollment_no, mobile, program.program_id, trust_id=trust_id)

            st_q = select(Student).filter_by(enrollment_no=enrollment_no)
            if trust_id:
                st_q = st_q.filter(Student.trust_id_fk == int(trust_id))
            student = db.session.execute(st_q).scalars().first()
            student_was_existing = student is not None
            if not student:
                student = Student(
                    enrollment_no=enrollment_no,
                    user_id_fk=user_id,
                    program_id_fk=program.program_id,
                    division_id_fk=(division.division_id if division else None),
                    trust_id_fk=trust_id,
                    last_name=surname,
                    first_name=student_name,
                    father_name=father_name,
                    mobile=mobile,
                    date_of_birth=dob,
                    gender=gender,
                    medium_tag=None,
                    photo_url=photo_url,
                    permanent_address=permanent_address,
                    current_semester=current_semester,
                    admission_academic_year=admission_academic_year or None,
                    intake_batch_id_fk=(intake_batch.intake_batch_id if intake_batch else None),
                    roll_no=roll_no,
                    aadhar_no=aadhar_no or None,
                    category=category or None,
                    is_active=True,
                )
                db.session.add(student)
                created += 1
            else:
                try:
                    student.is_active = True
                except Exception:
                    pass
                try:
                    if trust_id and hasattr(student, "trust_id_fk"):
                        student.trust_id_fk = trust_id
                except Exception:
                    pass
                student.program_id_fk = program.program_id
                student.division_id_fk = division.division_id if division else None
                if not student.user_id_fk and user_id:
                    student.user_id_fk = user_id
                student.last_name = surname or student.last_name
                student.first_name = student_name or student.first_name
                student.mobile = mobile or student.mobile
                student.father_name = father_name or student.father_name
                student.date_of_birth = dob or student.date_of_birth
                student.gender = gender or student.gender
                student.photo_url = photo_url or student.photo_url
                student.permanent_address = permanent_address or student.permanent_address
                student.current_semester = current_semester or student.current_semester
                if admission_academic_year:
                    student.admission_academic_year = admission_academic_year
                    student.intake_batch_id_fk = intake_batch.intake_batch_id if intake_batch else None
                if aadhar_no:
                    student.aadhar_no = aadhar_no
                if category:
                    student.category = category
                if roll_no:
                    student.roll_no = roll_no
                updated += 1
            # Assign medium with BCom defaulting to General when absent
            try:
                student.medium_tag = medium_tag or (student.medium_tag or None)
            except Exception:
                student.medium_tag = medium_tag or (student.medium_tag or None)
                errors.append(f"Row {row_idx}: failed to compute medium_tag due to data format")
            add_preview(row_idx, data, "update" if student_was_existing else "create")

    except InvalidFileException:
        # Fallback path: xlrd for legacy .xls or malformed files
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        colmap = normalize_headers(headers)

        for r in range(1, sheet.nrows):
            data = {}
            for c in range(sheet.ncols):
                key = colmap.get(c)
                if not key:
                    continue
                val = sheet.cell_value(r, c)
                # Convert Excel serial date
                if key == "date_of_birth" and sheet.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, book.datemode).date()
                    except Exception:
                        pass
                data[key] = val

            enrollment_no = cell_to_str(data.get("enrollment_no"))
            if not enrollment_no:
                try:
                    if not any(cell_to_str(v) for v in data.values()):
                        continue
                except Exception:
                    pass
                skipped += 1
                errors.append(f"Row {r+1}: missing enrollment_no; skipped")
                add_preview(r + 1, data, "skip", "error", "Missing enrollment number.")
                continue
            
            row_semester = to_int(data.get("current_semester"))
            if row_semester and row_semester != semester:
                skipped += 1
                errors.append(f"Row {r+1}: semester {row_semester} conflicts with selected semester {semester}; skipped")
                add_preview(r + 1, data, "skip", "error", f"Semester {row_semester} conflicts with selected semester {semester}.")
                continue
            row_admission_year_raw = cell_to_str(data.get("admission_academic_year")).strip()
            row_admission_year = normalize_academic_year(row_admission_year_raw)
            if row_admission_year and row_admission_year != admission_academic_year_hint:
                skipped += 1
                errors.append(f"Row {r+1}: admission year '{row_admission_year_raw}' conflicts with selected '{admission_academic_year_hint}'; skipped")
                add_preview(r + 1, data, "skip", "error", f"Admission year {row_admission_year_raw} conflicts with selected {admission_academic_year_hint}.")
                continue
            processed_enrollments.add(enrollment_no)

            row_division_code = cell_to_str(data.get("division_code")).strip().upper()
            if row_division_code and division_hint and row_division_code != division_hint:
                skipped += 1
                errors.append(f"Row {r+1}: division '{row_division_code}' conflicts with selected '{division_hint}'; skipped")
                add_preview(r + 1, data, "skip", "error", f"Division {row_division_code} conflicts with selected {division_hint}.")
                continue
            division_code = row_division_code or division_hint
            division = db.session.execute(select(Division).filter_by(program_id_fk=program.program_id, semester=semester, division_code=division_code)).scalars().first() if division_code else None
            if division_code and not division:
                plan = db.session.execute(select(ProgramDivisionPlan).filter_by(program_id_fk=program.program_id, semester=semester)).scalars().first()
                cap = None
                if plan:
                    try:
                        cap = int(plan.capacity_per_division)
                    except Exception:
                        cap = None
                if cap is None:
                    cap = 67 if (program.program_name or "").upper() == "BCA" else (Division.capacity.default.arg if hasattr(Division.capacity, 'default') else 60)
                division = Division(program_id_fk=program.program_id, semester=semester, division_code=division_code, capacity=cap)
                db.session.add(division)
                db.session.flush()
                divisions_created += 1

            surname = cell_to_str(data.get("last_name"))
            student_name = cell_to_str(data.get("first_name"))
            mobile = cell_to_str(data.get("mobile"))
            father_name = cell_to_str(data.get("father_name"))
            gender = cell_to_str(data.get("gender")).capitalize()
            if gender not in ("Male", "Female", "Other", ""):
                gender = ""
            photo_url = cell_to_str(data.get("photo_url"))
            permanent_address = cell_to_str(data.get("permanent_address"))
            # Optional medium parsing
            row_medium_raw = cell_to_str(data.get("medium_tag")).strip()
            if row_medium_raw and medium_hint and row_medium_raw.lower() != medium_hint.lower():
                skipped += 1
                errors.append(f"Row {r+1}: MOI '{row_medium_raw}' conflicts with selected '{medium_hint}'; skipped")
                add_preview(r + 1, data, "skip", "error", f"MOI {row_medium_raw} conflicts with selected {medium_hint}.")
                continue
            medium_raw = (row_medium_raw or medium_hint).lower()
            medium_map = {
                "": "",
                "general": "General",
                "eng": "English",
                "english": "English",
                "e": "English",
                "guj": "Gujarati",
                "gujarati": "Gujarati",
                "g": "Gujarati",
            }
            medium_tag = medium_map.get(medium_raw, "")
            allowed = []
            default_m = ""
            try:
                cfg_row = cfg.get(program.program_name or "") or {}
                allowed = cfg_row.get("mediums") or []
                default_list = cfg_row.get("default") or []
                default_m = default_list[0] if default_list else ""
            except Exception:
                allowed = []
                default_m = ""
            if not medium_tag:
                medium_tag = default_m
            if allowed and medium_tag and medium_tag not in allowed:
                errors.append(f"Row {r+1}: medium '{medium_tag}' not allowed for {program.program_name}")
                medium_tag = default_m
            mediums_seen.add((medium_tag or "").strip())

            dob_val = data.get("date_of_birth")
            dob = None
            if isinstance(dob_val, datetime):
                dob = dob_val
            elif isinstance(dob_val, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        dob = datetime.strptime(dob_val.strip(), fmt).date()
                        break
                    except Exception:
                        pass

            current_semester = semester
            admission_academic_year = admission_academic_year_hint
            intake_batch = selected_intake_batch if import_mode == "new" else None
            aadhar_no = cell_to_str(data.get("aadhar_no"))
            category = cell_to_str(data.get("category"))

            # Ensure User exists
            user_id = ensure_student_user(enrollment_no, mobile, program.program_id, trust_id=trust_id)

            st_q = select(Student).filter_by(enrollment_no=enrollment_no)
            if trust_id:
                st_q = st_q.filter(Student.trust_id_fk == int(trust_id))
            student = db.session.execute(st_q).scalars().first()
            student_was_existing = student is not None
            if not student:
                preview_action = "create"
                student = Student(
                    enrollment_no=enrollment_no,
                    user_id_fk=user_id,
                    program_id_fk=program.program_id,
                    division_id_fk=(division.division_id if division else None),
                    trust_id_fk=trust_id,
                    last_name=surname,
                    first_name=student_name,
                    father_name=father_name,
                    mobile=mobile,
                    date_of_birth=dob,
                    gender=gender,
                    medium_tag=None,
                    photo_url=photo_url,
                    permanent_address=permanent_address,
                    current_semester=current_semester,
                    admission_academic_year=admission_academic_year or None,
                    intake_batch_id_fk=(intake_batch.intake_batch_id if intake_batch else None),
                    aadhar_no=aadhar_no or None,
                    category=category or None,
                    is_active=True,
                )
                db.session.add(student)
                created += 1
            else:
                preview_action = "update"
                try:
                    student.is_active = True
                except Exception:
                    pass
                student.program_id_fk = program.program_id
                student.division_id_fk = division.division_id if division else None
                if not student.user_id_fk and user_id:
                    student.user_id_fk = user_id
                student.last_name = surname or student.last_name
                student.first_name = student_name or student.first_name
                student.mobile = mobile or student.mobile
                student.father_name = father_name or student.father_name
                student.date_of_birth = dob or student.date_of_birth
                student.gender = gender or student.gender
                student.photo_url = photo_url or student.photo_url
                student.permanent_address = permanent_address or student.permanent_address
                student.current_semester = current_semester or student.current_semester
                if admission_academic_year:
                    student.admission_academic_year = admission_academic_year
                    student.intake_batch_id_fk = intake_batch.intake_batch_id if intake_batch else None
                updated += 1
            # Assign medium with BCom defaulting to General when absent
            try:
                student.medium_tag = medium_tag or (student.medium_tag or None)
            except Exception:
                student.medium_tag = medium_tag or (student.medium_tag or None)
                errors.append(f"Row {r+1}: failed to compute medium_tag due to data format")
            add_preview(r + 1, data, preview_action)

    # Delete students that are in DB for this Program+Semester but NOT in the Excel file.
    # For multi-medium programs like B.Com, only delete students that share the same medium(s)
    # as rows present in this import, so importing English does not delete Gujarati, and vice versa.
    if replace_scope and mediums_seen:
        normalized_mediums = {m.strip() for m in mediums_seen}
        for existing in existing_students:
            enr = existing.enrollment_no
            if not enr or enr in processed_enrollments:
                continue
            existing_medium = (existing.medium_tag or "").strip()
            if existing_medium in normalized_mediums:
                db.session.delete(existing)
                deleted += 1
    elif replace_scope:
        # Fallback: no medium information, keep legacy behavior (program+semester full replacement)
        existing_enrollments = {s.enrollment_no for s in existing_students}
        students_to_delete = existing_enrollments - processed_enrollments
        for enr in students_to_delete:
            if enr:
                s_to_del = db.session.execute(select(Student).filter_by(enrollment_no=enr)).scalars().first()
                if s_to_del:
                    db.session.delete(s_to_del)
                    deleted += 1

    if import_mode == "new" and selected_intake_batch:
        mapped_count = db.session.execute(
            select(func.count()).select_from(Student).filter_by(
                intake_batch_id_fk=selected_intake_batch.intake_batch_id,
                is_active=True,
            )
        ).scalar() or 0
        if mapped_count > int(selected_intake_batch.approved_intake or 0):
            db.session.rollback()
            raise ValueError(
                f"Approved intake exceeded: {mapped_count} active students would be linked to an "
                f"approved intake of {selected_intake_batch.approved_intake}."
            )

    if not dry_run:
        db.session.commit()
    else:
        db.session.rollback()
    print(f"Imported from {path}: created={created}, updated={updated}, skipped={skipped}, deleted={deleted}, divisions_created={divisions_created}, errors={len(errors)}")
    # Return a detailed report for UI display
    start_year = int(admission_academic_year_hint[:4])
    completion_start = start_year + max(int(program.program_duration_years or 1), 1)
    completion_academic_year = f"{completion_start}-{str(completion_start + 1)[-2:]}"
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "deleted": deleted,
        "errors_count": len(errors),
        "errors": errors,
        "divisions_created": divisions_created,
        "program_name": program.program_name,
        "program_id": program.program_id,
        "semester": semester,
        "import_mode": import_mode,
        "admission_academic_year": admission_academic_year_hint,
        "completion_academic_year": completion_academic_year,
        "medium_tag": medium_hint or None,
        "division_code": division_hint or None,
        "approved_intake": (selected_intake_batch.approved_intake if selected_intake_batch else None),
        "preview_rows": preview_rows,
        "preview_total": preview_total,
        "preview_limit": preview_limit,
        "path": path,
    }


def main():
    app = create_app()
    with app.app_context():
        args = sys.argv[1:]
        if not args:
            # Default to BCom bulk import files when no args provided
            args = [
                r"c:\project\CMSv5\B.Com\Bulk BCom Student Import for Semester 3.xlsx",
                r"c:\project\CMSv5\B.Com\Bulk BCom Student Import for Semester 5.xlsx",
            ]
        for p in args:
            # Detect program from filename and use semester hints from filenames
            program_name = detect_program_from_filename(p)
            sem_hint = find_semester_from_filename(p)
            import_excel(p, program_name=program_name, semester_hint=sem_hint or None)


if __name__ == "__main__":
    main()
