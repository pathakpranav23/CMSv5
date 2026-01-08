import os
import sys
import re
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Ensure project root is on sys.path when running from scripts/
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from cms_app import create_app, db
from cms_app.models import Program, FeeStructure


def cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    try:
        return str(v).strip()
    except Exception:
        return ""


def normalize_program_name(name: str) -> str:
    n = (name or "").strip()
    # Common cleanup
    n = re.sub(r"\s+", " ", n)
    # Consistent casing for comparisons
    return n.upper()


def _import_wide_sheet(ws) -> Tuple[int, int]:
    """Handle wide-format sheet: columns after Description are program names."""
    # Detect header row
    header_row_idx = None
    programs: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [cell_to_str(c) for c in row]
        if len(cells) >= 2 and cells[0].upper() in ("SR NO", "SRNO", "S.NO") and cells[1].upper() == "DESCRIPTION":
            header_row_idx = i
            programs = [normalize_program_name(cells[j]) for j in range(2, len(cells))]
            break
    if header_row_idx is None:
        raise RuntimeError("Could not locate header row with 'Sr No' and 'Description'.")

    # Ensure programs exist
    program_objs: Dict[str, Program] = {}
    for p_name_raw in programs:
        if not p_name_raw:
            continue
        p_lookup = p_name_raw.replace(".", "").replace("(E)", "E").replace("(G)", "G").strip()
        prog = Program.query.filter(Program.program_name.ilike(p_lookup)).first()
        if not prog:
            prog = Program(program_name=p_lookup, program_duration_years=3)
            db.session.add(prog)
            db.session.flush()
        program_objs[p_name_raw] = prog

    created = 0
    updated = 0

    # Semester detection
    current_semester: Optional[int] = 1
    semester_hint_patterns = [
        re.compile(r"\bSEM(?:ESTER)?\s*1\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*2\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*3\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*4\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*5\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*6\b", re.IGNORECASE),
    ]

    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        cells = [cell_to_str(c) for c in row]
        if len(cells) < 3:
            continue

        desc_upper = (cells[1] or "").upper()
        # Skip banners and TOTAL rows
        if desc_upper in ("FEE PATRAK- 2024 ALL UNIVERSITY AFFILIATED COURSES", "M.K.BHAVNAGAR UNIVERSITY FEE PATRAK"):
            continue
        if re.search(r"\bTOTAL\b", desc_upper):
            continue

        for s_idx, pat in enumerate(semester_hint_patterns, start=1):
            if pat.search(desc_upper):
                current_semester = s_idx
                break

        component = cells[1]
        if not component:
            continue

        # per-program amounts
        for j in range(2, len(cells)):
            prog_key = programs[j - 2]
            prog_obj = program_objs.get(prog_key)
            if not prog_obj:
                continue
            amt_raw = row[j]
            try:
                amount = float(amt_raw) if amt_raw is not None and str(amt_raw).strip() != "" else None
            except Exception:
                amount = None
            if amount is None:
                continue
            fs = FeeStructure.query.filter_by(program_id_fk=prog_obj.program_id, component_name=component, semester=current_semester).first()
            if not fs:
                fs = FeeStructure(program_id_fk=prog_obj.program_id, component_name=component, semester=current_semester, amount=amount)
                db.session.add(fs)
                created += 1
            else:
                fs.amount = amount
                updated += 1

    db.session.commit()

    # Update program duration years by observed max semester
    if current_semester is not None:
        for prog in program_objs.values():
            max_sem = db.session.query(db.func.max(FeeStructure.semester)).filter(FeeStructure.program_id_fk == prog.program_id).scalar()
            if max_sem and isinstance(max_sem, int):
                years = (max_sem + 1) // 2
                try:
                    prog.program_duration_years = max(1, years)
                except Exception:
                    pass
        db.session.commit()

    return created, updated


def _import_single_program_sheet(ws, program_name_hint: Optional[str]) -> Tuple[int, int]:
    """Handle sheet that contains one program fee structure vertically."""
    # Determine header row
    header_row_idx = None
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [cell_to_str(c) for c in row]
        if len(cells) >= 2 and cells[0].upper() in ("SR NO", "SRNO", "S.NO") and cells[1].upper() == "DESCRIPTION":
            header_row_idx = i
            headers = [cell_to_str(c) for c in row]
            break
    if header_row_idx is None:
        # Fallback: try first row as header if it looks like text labels
        first = [cell_to_str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if len(first) >= 2:
            header_row_idx = 1
            headers = first
        else:
            raise RuntimeError("Could not locate header row with 'Sr No' and 'Description'.")

    # Determine program name: from header third column or sheet title or hint
    program_name = None
    if len(headers) >= 3 and headers[2]:
        program_name = normalize_program_name(headers[2])
    if not program_name:
        program_name = normalize_program_name(program_name_hint or ws.title or "")
    if not program_name:
        raise RuntimeError("Program name could not be determined for sheet")

    # Ensure program exists
    p_lookup = program_name.replace(".", "").replace("(E)", "E").replace("(G)", "G").strip()
    prog = Program.query.filter(Program.program_name.ilike(p_lookup)).first()
    if not prog:
        prog = Program(program_name=p_lookup, program_duration_years=3)
        db.session.add(prog)
        db.session.flush()

    created = 0
    updated = 0

    # Find amount column index after Description
    amount_col_idx = None
    # Prefer headers like Amount/Total/Fee/Rupees
    for idx in range(2, len(headers)):
        h = headers[idx].upper()
        if h in ("AMOUNT", "TOTAL", "FEE", "RUPEES", "RS", "AMT"):
            amount_col_idx = idx
            break
    if amount_col_idx is None:
        # Infer by scanning a few rows for numeric cell
        scan_rows = list(ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 10, values_only=True))
        for idx in range(2, max(len(r) for r in scan_rows) if scan_rows else 3):
            for r in scan_rows:
                if idx < len(r):
                    try:
                        v = r[idx]
                        if v is not None and str(v).strip() != "" and float(v) == float(v):
                            amount_col_idx = idx
                            break
                    except Exception:
                        pass
            if amount_col_idx is not None:
                break
    if amount_col_idx is None:
        amount_col_idx = 2  # fallback

    # Semester detection
    current_semester: Optional[int] = 1
    semester_hint_patterns = [
        re.compile(r"\bSEM(?:ESTER)?\s*1\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*2\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*3\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*4\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*5\b", re.IGNORECASE),
        re.compile(r"\bSEM(?:ESTER)?\s*6\b", re.IGNORECASE),
    ]

    # Iterate data rows
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cells = [cell_to_str(c) for c in row]
        if len(cells) < 2:
            continue
        desc_upper = (cells[1] or "").upper()
        if desc_upper in ("FEE PATRAK- 2024 ALL UNIVERSITY AFFILIATED COURSES", "M.K.BHAVNAGAR UNIVERSITY FEE PATRAK"):
            continue
        if re.search(r"\bTOTAL\b", desc_upper):
            continue
        for s_idx, pat in enumerate(semester_hint_patterns, start=1):
            if pat.search(desc_upper):
                current_semester = s_idx
                break
        component = cells[1]
        if not component:
            continue
        amt_raw = row[amount_col_idx] if amount_col_idx < len(row) else None
        try:
            amount = float(amt_raw) if amt_raw is not None and str(amt_raw).strip() != "" else None
        except Exception:
            amount = None
        if amount is None:
            continue
        fs = FeeStructure.query.filter_by(program_id_fk=prog.program_id, component_name=component, semester=current_semester).first()
        if not fs:
            fs = FeeStructure(program_id_fk=prog.program_id, component_name=component, semester=current_semester, amount=amount)
            db.session.add(fs)
            created += 1
        else:
            fs.amount = amount
            updated += 1

    db.session.commit()

    # Update duration by observed max semester
    if current_semester is not None:
        max_sem = db.session.query(db.func.max(FeeStructure.semester)).filter(FeeStructure.program_id_fk == prog.program_id).scalar()
        if max_sem and isinstance(max_sem, int):
            years = (max_sem + 1) // 2
            try:
                prog.program_duration_years = max(1, years)
            except Exception:
                pass
        db.session.commit()

    return created, updated


def import_fee_structure(path: str) -> Tuple[int, int]:
    """Import fee structures from an Excel that may contain multiple sheets.

    - If a sheet is wide-format (programs across columns), use _import_wide_sheet.
    - Else treat it as single-program vertical format and use _import_single_program_sheet.
    """
    wb = load_workbook(filename=path, data_only=True)
    total_created = 0
    total_updated = 0
    for ws in wb.worksheets:
        # Heuristic: check if header row has programs across columns
        is_wide = False
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [cell_to_str(c) for c in row]
            if len(cells) >= 2 and cells[0].upper() in ("SR NO", "SRNO", "S.NO") and cells[1].upper() == "DESCRIPTION":
                # consider wide if there are at least one non-empty header after col 2
                after = [cell_to_str(c) for c in cells[2:]]
                non_empty_headers = [h for h in after if h]
                is_wide = len(non_empty_headers) >= 1
                break
        try:
            if is_wide:
                c, u = _import_wide_sheet(ws)
            else:
                c, u = _import_single_program_sheet(ws, ws.title)
            total_created += c
            total_updated += u
        except Exception as e:
            # Continue with other sheets but report the issue
            print(f"Sheet '{ws.title}': skipped due to error: {e}")
            continue

    return total_created, total_updated


def main():
    # Allow passing a custom Excel path; default to project root sample
    path = None
    if len(sys.argv) >= 2:
        path = sys.argv[1]
    if not path:
        path = os.path.join(BASE_DIR, "Fee Structure for SBPET.xlsx")
    app = create_app()
    with app.app_context():
        created, updated = import_fee_structure(path)
        print(f"Fee structure import completed. created={created}, updated={updated}")


if __name__ == "__main__":
    main()