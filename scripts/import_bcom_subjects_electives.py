import os
import re
import sys
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Ensure project root is on sys.path when running from scripts/
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from cms_app import create_app, db
from cms_app.models import Program, SubjectType, Subject, CreditStructure


# Synonyms to normalize headers from university Excel files
HEADER_MAP: Dict[str, List[str]] = {
    "subject_name": [
        "subject",
        "subject name",
        "course",
        "paper",
        "title",
        "name",
        "sub name",
    ],
    "subject_code": [
        "code",
        "subject code",
        "course code",
        "code no",
        "subject id",
        "sub code",
    ],
    "paper_code": [
        "paper code",
        "paper id",
        "paper no",
    ],
    "subject_type": [
        "type",
        "subject type",
        "category",
        "course type",
        "type code",
    ],
    "semester": [
        "semester",
        "sem",
        "semester no",
        "sem no",
    ],
    "theory_credits": [
        "theory credits",
        "theory",
        "th credits",
        "credits theory",
        "th",
        "credit",
    ],
    "practical_credits": [
        "practical credits",
        "practical",
        "pr credits",
        "credits practical",
        "pr",
    ],
    "total_credits": [
        "total credits",
        "credits",
        "credit points",
        "credits total",
        "points",
    ],
    # Elective and offering metadata
    "is_elective": [
        "elective",
        "is elective",
        "elective course",
        "elective?",
    ],
    "offered_programs": [
        "offered to",
        "offered for program",
        "for program",
        "programme offered to",
        "program offered to",
        "given program",
        "target program",
        "offered for",
        "offered in",
        "program",
    ],
    "capacity": [
        "capacity",
        "seat limit",
        "max seats",
        "intake",
    ],
    "elective_group_id": [
        "elective group",
        "group",
        "cluster",
        "bucket",
        "group id",
    ],
    "medium_tag": [
        "medium",
        "language",
        "medium of instruction",
    ],
}


def normalize_headers(headers: List[Any]) -> Dict[int, str]:
    colmap: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        key = str(h or "").strip().lower()
        if not key:
            continue
        for target, synonyms in HEADER_MAP.items():
            if key == target:
                colmap[idx] = target
                break
            if key in synonyms:
                colmap[idx] = target
                break
    return colmap


def cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if abs(v - int(v)) < 1e-9:
            return str(int(v))
        return str(v)
    return str(v)


def detect_semester_from_path(path: str) -> int:
    s = (path or "").lower().replace("_", " ").replace("-", " ")
    m_sem = re.search(r"\b(?:sem(?:ester)?)\b[\s_-]*([0-9]{1,2})", s, flags=re.IGNORECASE)
    if not m_sem:
        m_sem = re.search(r"\bs[\s_-]*([0-9]{1,2})", s)
    try:
        return int(m_sem.group(1)) if m_sem else 1
    except Exception:
        return 1


def get_or_create_subject_type(code: str) -> SubjectType:
    code_norm = (code or "MAJOR").strip().upper()
    st = SubjectType.query.filter_by(type_code=code_norm).first()
    if not st:
        st = SubjectType(type_code=code_norm, description=None)
        db.session.add(st)
        db.session.flush()
    return st


def normalize_program_name(raw: str) -> str:
    s = (raw or "").strip().lower()
    if not s:
        return ""
    if ("bcom" in s) or ("b.com" in s) or ("b com" in s):
        return "BCOM"
    if "bca" in s:
        return "BCA"
    if "bba" in s:
        return "BBA"
    return raw.strip().upper()


def split_programs(raw: str) -> List[str]:
    s = (raw or "").strip()
    if not s:
        return []
    tokens = re.split(r"[,/&|]+|\band\b", s, flags=re.IGNORECASE)
    return [normalize_program_name(t) for t in tokens if normalize_program_name(t)]


def parse_bool(v: Any) -> bool:
    s = cell_to_str(v).strip().lower()
    if not s:
        return False
    return s in {"1", "true", "t", "yes", "y"}


def derive_is_elective(subject_type_code: str, elective_raw: Any) -> bool:
    if parse_bool(elective_raw):
        return True
    code = (subject_type_code or "").strip().upper()
    elective_like = {"ELECTIVE", "SEC", "AEC", "GE", "DSE", "OE"}
    for token in elective_like:
        if token in code:
            return True
    return False


def upsert_subject_for_program(
    program_name: str,
    subject_name: str,
    subject_code: str,
    paper_code: str,
    subject_type_code: str,
    semester: int,
    th: int,
    pr: int,
    total: int,
    is_elective_flag: bool,
    capacity_val: str,
    group_id: str,
    medium_tag: str,
) -> Tuple[Subject, bool]:
    program = Program.query.filter_by(program_name=program_name).first()
    if not program:
        program = Program(program_name=program_name, program_duration_years=3)
        db.session.add(program)
        db.session.flush()

    st = get_or_create_subject_type(subject_type_code)
    subj = None
    if subject_code:
        subj = Subject.query.filter_by(program_id_fk=program.program_id, subject_code=subject_code).first()
    if not subj:
        subj = Subject.query.filter_by(program_id_fk=program.program_id, subject_name=subject_name, semester=semester).first()

    created = False
    if not subj:
        subj = Subject(
            program_id_fk=program.program_id,
            subject_type_id_fk=st.type_id,
            subject_name=subject_name,
            subject_code=(subject_code or None),
            paper_code=(paper_code or None),
            semester=semester,
            is_elective=is_elective_flag,
            capacity=int(float(capacity_val)) if capacity_val else None,
            elective_group_id=(group_id or None),
            medium_tag=(medium_tag or None),
        )
        db.session.add(subj)
        db.session.flush()
        created = True
    else:
        subj.subject_type_id_fk = st.type_id
        if subject_code:
            subj.subject_code = subject_code
        if paper_code:
            subj.paper_code = paper_code
        subj.is_elective = is_elective_flag
        if capacity_val:
            try:
                subj.capacity = int(float(capacity_val))
            except Exception:
                pass
        subj.elective_group_id = (group_id or None)
        if medium_tag:
            subj.medium_tag = medium_tag

    cs = CreditStructure.query.filter_by(subject_id_fk=subj.subject_id).first()
    if not cs:
        cs = CreditStructure(subject_id_fk=subj.subject_id, theory_credits=th, practical_credits=pr, total_credits=total)
        db.session.add(cs)
    else:
        cs.theory_credits = th
        cs.practical_credits = pr
        cs.total_credits = total

    return subj, created


def import_subjects_for_bcom(path: str) -> Tuple[int, int, int]:
    default_semester = detect_semester_from_path(path)
    created = 0
    updated = 0
    offered_created = 0

    try:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        colmap = normalize_headers(headers)

        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            data = {}
            for idx, key in colmap.items():
                data[key] = values[idx] if idx < len(values) else None

            subject_name = cell_to_str(data.get("subject_name")).strip()
            subject_code = cell_to_str(data.get("subject_code")).strip()
            paper_code = cell_to_str(data.get("paper_code")).strip()
            subject_type_code = cell_to_str(data.get("subject_type")).strip().upper() or "MAJOR"
            sem_val = cell_to_str(data.get("semester")).strip()
            semester = default_semester
            if sem_val.isdigit():
                try:
                    semester = int(sem_val)
                except Exception:
                    semester = default_semester

            if not subject_name:
                continue

            is_elective_flag = derive_is_elective(subject_type_code, data.get("is_elective"))
            offered_programs = split_programs(cell_to_str(data.get("offered_programs")))
            capacity_val = cell_to_str(data.get("capacity")).strip()
            group_id = cell_to_str(data.get("elective_group_id")).strip() or None
            medium_tag = cell_to_str(data.get("medium_tag")).strip() or None

            def to_int_safe(x: Any) -> int:
                s = cell_to_str(x).strip()
                if not s:
                    return 0
                try:
                    return int(float(s))
                except Exception:
                    return 0

            th = to_int_safe(data.get("theory_credits"))
            pr = to_int_safe(data.get("practical_credits"))
            total = to_int_safe(data.get("total_credits"))
            if not total:
                total = th + pr

            subj, was_created = upsert_subject_for_program(
                "BCOM",
                subject_name,
                subject_code,
                paper_code,
                subject_type_code,
                semester,
                th,
                pr,
                total,
                is_elective_flag,
                capacity_val,
                group_id,
                medium_tag,
            )
            if was_created:
                created += 1
            else:
                updated += 1

            # Offer in additional programs as electives
            for offered_prog in offered_programs:
                if offered_prog == "BCOM":
                    continue
                _, offer_created = upsert_subject_for_program(
                    offered_prog,
                    subject_name,
                    subject_code,
                    paper_code,
                    subject_type_code,
                    semester,
                    th,
                    pr,
                    total,
                    True,  # force elective in offered program
                    capacity_val,
                    group_id,
                    medium_tag,
                )
                if offer_created:
                    offered_created += 1

    except InvalidFileException:
        raise RuntimeError(f"Legacy .xls is not supported for path: {path}")

    db.session.commit()
    print(
        f"Imported BCom subjects from {path}: created={created}, updated={updated}, offered_created={offered_created}"
    )
    return created, updated, offered_created


def main():
    app = create_app()
    with app.app_context():
        args = sys.argv[1:]
        if not args:
            args = [
                r"c:\project\CMSv5\B.Com\BCom Semester 3 Subject Detail V2.xlsx",
                r"c:\project\CMSv5\B.Com\BCom Semester 5 Subject Detail V2.xlsx",
            ]
        for p in args:
            print(f"Processing BCom subjects: default_semester={detect_semester_from_path(p)}, path={p}")
            import_subjects_for_bcom(p)


if __name__ == "__main__":
    main()