import os
import sys

from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from cms_app import create_app, db  # noqa: E402
from cms_app.models import Program, Student  # noqa: E402
from sqlalchemy import select, func  # noqa: E402
from scripts.import_students import normalize_headers, cell_to_str, HEADER_MAP  # noqa: E402


def load_enrollment_category_mapping(path: str):
    if not os.path.exists(path):
        print("File not found:", path)
        return {}
    wb = load_workbook(path, data_only=True)
    mapping = {}
    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [h if h is not None else "" for h in header_row]
        colmap = normalize_headers(headers)
        enr_idx = None
        cat_idx = None
        for idx, key in colmap.items():
            if key == "enrollment_no":
                enr_idx = idx
            elif key == "category":
                cat_idx = idx
        if enr_idx is None or cat_idx is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row)
            if enr_idx >= len(values):
                continue
            enr = cell_to_str(values[enr_idx])
            if not enr:
                continue
            cat_val = ""
            if cat_idx < len(values):
                cat_val = cell_to_str(values[cat_idx])
            if not cat_val:
                continue
            mapping[enr] = cat_val
    return mapping


def main():
    excel_path = r"c:\project\CMSv5\DATA FOR IMPORT EXPORT\BCA\BCA Sem 4 Bulk Student Data 2026 22 feb.xlsx"
    mapping = load_enrollment_category_mapping(excel_path)
    print("Loaded category entries from Excel:", len(mapping))
    app = create_app()
    with app.app_context():
        program = db.session.execute(
            select(Program).filter(func.lower(Program.program_name) == "bca")
        ).scalars().first()
        if not program:
            print("No BCA program found")
            return
        q = (
            select(Student)
            .filter(Student.program_id_fk == program.program_id)
            .filter(Student.current_semester == 4)
            .filter((Student.category == None) | (func.trim(Student.category) == ""))
        )
        students = db.session.execute(q).scalars().all()
        print("BCA Sem 4 students with missing category before update:", len(students))
        updated = 0
        for s in students:
            enr = cell_to_str(getattr(s, "enrollment_no", ""))
            cat = mapping.get(enr)
            if not cat:
                continue
            s.category = cat
            updated += 1
        if updated:
            db.session.commit()
        print("Updated students:", updated)


if __name__ == "__main__":
    main()

