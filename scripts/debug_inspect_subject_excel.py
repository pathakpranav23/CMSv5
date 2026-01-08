import os
import sys
from typing import Any

# Ensure project root
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from openpyxl import load_workbook
from cms_app import create_app
from scripts.import_subjects import normalize_headers


def inspect_headers(path: str):
    print(f"Inspecting: {path}")
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print("Raw headers:")
    for i, h in enumerate(headers):
        print(f"  [{i}] {repr(h)}")
    colmap = normalize_headers(headers)
    print("Normalized mapping (index -> key):")
    for idx, key in sorted(colmap.items()):
        print(f"  {idx} -> {key}")
    # Show first 10 data rows
    print("\nSample rows (first 10):")
    for r_i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        values = [cell.value for cell in row]
        data = {}
        for idx, key in colmap.items():
            if idx < len(values):
                data[key] = values[idx]
        subj = (str(data.get("subject_name")) if data.get("subject_name") is not None else "").strip()
        if r_i <= 10:
            print(f"  Row {r_i}: subject_name={repr(subj)}, subject_code={repr(data.get('subject_code'))}, paper_code={repr(data.get('paper_code'))}, semester={repr(data.get('semester'))}, type={repr(data.get('subject_type'))}")
        else:
            break


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        inspect_headers(r"c:\project\CMSv5\B.Com\Bcom English Semester 3 Subject Detail.xlsx")