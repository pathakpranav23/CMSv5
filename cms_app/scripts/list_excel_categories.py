from openpyxl import load_workbook
import os


def main():
    path = r"c:\project\CMSv5\DATA FOR IMPORT EXPORT\BCA\BCA Sem 2 Bulk Student Data 2026.xlsx"
    if not os.path.exists(path):
        print("FILE_NOT_FOUND", path)
        return
    wb = load_workbook(path, data_only=True)
    any_found = False
    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        category_cols = []
        for idx, val in enumerate(header_row, start=1):
            if not val:
                continue
            text = str(val).strip().lower()
            if "category" in text or "caste" in text:
                category_cols.append((idx, val))
        if not category_cols:
            continue
        any_found = True
        print(f"SHEET: {ws.title}")
        for col_idx, col_header in category_cols:
            values = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if col_idx - 1 >= len(row):
                    continue
                v = row[col_idx - 1]
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                values.add(s)
            print(f"  COLUMN: {col_header}")
            for v in sorted(values):
                print(f"    - {v}")
    if not any_found:
        print("NO_CATEGORY_COLUMNS_FOUND")


if __name__ == "__main__":
    main()

